"""데이터 접근 계층.

dataset_2(반도체 DS 수율·공정설계 & DX 스마트팩토리 데이터 레이크, 차원 3 + 팩트 3)를
읽어, FDC/수율/KPI/통합 에이전트가 공통으로 쓰는 조회·조인 함수를 제공한다. MCP
서버(mymcp/server.py)는 이 모듈의 함수만 호출한다 — 데이터 로직과 프로토콜 계층을
분리해두면 나중에 실제 MCP SDK로 전송 계층만 바꿔도 이 파일은 그대로 재사용할 수 있다.
"""
from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import pandas as pd

from insight_agent.config import CAUSAL_YIELD_DEFICIT_PP, DATASET_DIR

TABLE_FILES = {
    "dim_ai_automation_role": "dim_ai_automation_role.csv",
    "dim_process_design": "dim_process_design.csv",
    "dim_semicon_dx_architecture": "dim_semicon_dx_architecture.csv",
    "fact_wafer_lot_yield": "fact_wafer_lot_yield.csv",
    "fact_fdc_chamber_sensor": "fact_fdc_chamber_sensor.csv",
    "fact_dx_smart_factory_kpi": "fact_dx_smart_factory_kpi.csv",
}


@dataclass
class Tables:
    dim_ai_automation_role: pd.DataFrame
    dim_process_design: pd.DataFrame
    dim_semicon_dx_architecture: pd.DataFrame
    fact_wafer_lot_yield: pd.DataFrame
    fact_fdc_chamber_sensor: pd.DataFrame
    fact_dx_smart_factory_kpi: pd.DataFrame


def load_tables(data_dir: Path | str = DATASET_DIR) -> Tables:
    data_dir = Path(data_dir)
    frames = {
        name: pd.read_csv(data_dir / filename, encoding="utf-8-sig")
        for name, filename in TABLE_FILES.items()
    }
    return Tables(**frames)


def df_to_records(df: pd.DataFrame) -> list[dict]:
    """DataFrame -> JSON-safe list[dict] (numpy int64/float64를 안전하게 변환)."""
    if df.empty:
        return []
    return json.loads(df.to_json(orient="records", force_ascii=False))


def get_fdc_anomalies(
    tables: Tables,
    process_id: Optional[str] = None,
    chamber_id: Optional[str] = None,
    max_vm_error_pct: float = 1.5,
) -> pd.DataFrame:
    """인터록(fdc_interlock_flag)이 발생했거나, 가상계측(VM) 예측 두께와 실측 두께의
    오차율이 임계치를 넘는 FDC 센서 로그. 공정명/제어 AI 에이전트명을 조인해 반환한다."""
    fdc = tables.fact_fdc_chamber_sensor.copy()
    vm_error_pct = (fdc["actual_thickness_nm"] - fdc["vm_pred_thickness_nm"]).abs() / fdc[
        "actual_thickness_nm"
    ] * 100
    fdc["vm_error_pct"] = vm_error_pct.round(3)
    mask = (fdc["fdc_interlock_flag"] == 1) | (vm_error_pct > max_vm_error_pct)
    if process_id:
        mask &= fdc["process_id"] == process_id
    if chamber_id:
        mask &= fdc["chamber_id"] == chamber_id
    result = fdc.loc[mask].merge(
        tables.dim_process_design[["process_id", "process_name"]], on="process_id", how="left"
    ).merge(
        tables.dim_ai_automation_role[["agent_id", "agent_name"]],
        left_on="controlling_ai_agent",
        right_on="agent_id",
        how="left",
    )
    return result.sort_values("timestamp", ascending=False)


def get_yield_defects(
    tables: Tables,
    defect_mechanism: Optional[str] = None,
    product_node: Optional[str] = None,
) -> pd.DataFrame:
    """웨이퍼 로트 수율 이력을 결함 메커니즘/공정 노드로 필터링. 기본값은 결함이 있는
    (major_defect_mechanism != 'None(Clean)') 로트만 반환한다."""
    lots = tables.fact_wafer_lot_yield
    if defect_mechanism:
        lots = lots[lots["major_defect_mechanism"] == defect_mechanism]
    else:
        lots = lots[lots["major_defect_mechanism"] != "None(Clean)"]
    if product_node:
        lots = lots[lots["product_node"] == product_node]
    return lots.sort_values("production_date", ascending=False)


def get_dx_kpi_trend(
    tables: Tables,
    year_month: Optional[str] = None,
    quarter: Optional[str] = None,
) -> pd.DataFrame:
    """DX 스마트팩토리 월별 운영 성과(KPI) 추이를 연월/분기로 필터링해 조회한다."""
    kpi = tables.fact_dx_smart_factory_kpi
    if year_month:
        kpi = kpi[kpi["year_month"] == year_month]
    if quarter:
        kpi = kpi[kpi["quarter"] == quarter]
    return kpi.sort_values("year_month")


def build_causal_report(tables: Tables, lot_id: str) -> dict:
    """FDC 설비 이상 -> 웨이퍼 수율 -> 담당 AI 에이전트를 lot_id 기준으로 엮은 통합 리포트.

    이 프로젝트의 핵심 가치다: fact_fdc_chamber_sensor.lot_id로 팩트 테이블들이
    이어져 있어야만 "이 챔버 이상이 저 로트 수율 하락의 원인인가"를 한 번에 답할 수 있다.
    """
    lot_rows = tables.fact_wafer_lot_yield[tables.fact_wafer_lot_yield["lot_id"] == lot_id]
    if lot_rows.empty:
        raise ValueError(f"unknown lot_id: {lot_id}")
    lot = lot_rows.iloc[0]

    fdc = tables.fact_fdc_chamber_sensor[tables.fact_fdc_chamber_sensor["lot_id"] == lot_id].copy()
    vm_error_pct = (fdc["actual_thickness_nm"] - fdc["vm_pred_thickness_nm"]).abs() / fdc[
        "actual_thickness_nm"
    ] * 100 if not fdc.empty else pd.Series(dtype=float)

    # 인터록 플래그와 VM 오차 초과는 별개의 신호가 아니라 같은 행에 겹쳐 나타날 수
    # 있다 (dataset_2에서는 134건이 예외 없이 전부 일치한다 — 챔버 압력·플라즈마
    # 임피던스·서셉터 온도가 함께 튀는 하나의 물리 사건이 두 컬럼에 각각 기록된 것).
    # 두 값을 따로 세어 나란히 싣기만 하면 읽는 쪽이 독립된 두 근거로 오해하므로,
    # 중복을 제거한 행 수(fdc_anomaly_row_count)와 겹친 행 수를 함께 싣는다.
    interlock_mask = fdc["fdc_interlock_flag"] == 1 if not fdc.empty else pd.Series(dtype=bool)
    vm_over_mask = vm_error_pct > 1.5 if not fdc.empty else pd.Series(dtype=bool)

    processes = tables.dim_process_design[
        tables.dim_process_design["process_id"].isin(fdc["process_id"])
    ][["process_id", "process_name", "target_cpk"]]
    agents = tables.dim_ai_automation_role[
        tables.dim_ai_automation_role["agent_id"].isin(fdc["controlling_ai_agent"])
    ][["agent_id", "agent_name", "core_responsibility"]]

    same_node = tables.fact_wafer_lot_yield[
        tables.fact_wafer_lot_yield["product_node"] == lot["product_node"]
    ]

    year_month = str(lot["production_date"])[:7]
    kpi_context = tables.fact_dx_smart_factory_kpi[
        tables.fact_dx_smart_factory_kpi["year_month"] == year_month
    ]

    node_avg = round(float(same_node["die_yield_pct"].mean()), 2) if not same_node.empty else None
    anomaly_row_count = int((interlock_mask | vm_over_mask).sum()) if not fdc.empty else 0
    yield_delta_pp = (
        round(float(lot["die_yield_pct"]) - node_avg, 2) if node_avg is not None else None
    )

    return {
        "lot_id": lot_id,
        "production_date": str(lot["production_date"]),
        "product_node": lot["product_node"],
        "company_name": lot["company_name"],
        "business_unit": lot["business_unit"],
        "wafer_yield_pct": float(lot["wafer_yield_pct"]),
        "die_yield_pct": float(lot["die_yield_pct"]),
        "scrap_wafer_qty": int(lot["scrap_wafer_qty"]),
        "major_defect_mechanism": lot["major_defect_mechanism"],
        "fdc_row_count": int(len(fdc)),
        "fdc_interlock_count": int(interlock_mask.sum()) if not fdc.empty else 0,
        "fdc_vm_error_anomaly_count": int(vm_over_mask.sum()) if not fdc.empty else 0,
        # 위 두 값의 합집합/교집합. 합집합이 '실제로 이상이 있는 행 수'이고,
        # 교집합이 크면 두 값을 독립 근거로 셀 수 없다는 신호다.
        "fdc_anomaly_row_count": anomaly_row_count,
        "fdc_interlock_vm_coincident_count": (
            int((interlock_mask & vm_over_mask).sum()) if not fdc.empty else 0
        ),
        "involved_processes": df_to_records(processes),
        "controlling_agents": df_to_records(agents),
        "product_node_avg_die_yield_pct": node_avg,
        "die_yield_delta_pp": yield_delta_pp,
        # 리포트 이름이 'causal'이라고 해서 인과가 확정된 것이 아니다.
        # 판정은 항상 UNKNOWN이고, 사유가 무엇인지를 함께 싣는다.
        "causal_verdict": judge_causality(anomaly_row_count, yield_delta_pp),
        "dx_kpi_context": df_to_records(kpi_context)[0] if not kpi_context.empty else None,
    }


def judge_causality(anomaly_row_count: int, yield_delta_pp: Optional[float]) -> dict:
    """"FDC 이상이 이 로트의 수율을 떨어뜨렸는가"에 대한 판정.

    **이 함수는 인과를 확정하지 않는다.** 판정은 항상 UNKNOWN이고, 달라지는 것은
    "왜 근거가 없는지"뿐이다. 두 가지 이유에서다.

    1. 단일 로트 관측에는 비교군이 없다. 같은 조건에서 이상이 없었을 때의 수율을
       모르므로, 이상과 저수율이 함께 보여도 인과가 아니라 동시 발생일 뿐이다.
    2. dataset_2 전체에서도 로트당 인터록 건수와 다이 수율의 상관이 +0.005로
       사실상 0이다 (decisions.md D-004). 무결함 로트와 결함 로트의 평균 수율이
       같고, 인터록 로트가 오히려 근소하게 높다.

    상위 CLAUDE.md 근거 규율 2·3번을 그대로 따른다 -- "못 대면 UNKNOWN으로 남긴다",
    "UNKNOWN은 사유를 함께 기록한다. 왜 근거가 없는지가 근거만큼 중요하다."
    """
    base = {
        "verdict": "UNKNOWN",
        "anomaly_row_count": anomaly_row_count,
        "yield_delta_pp": yield_delta_pp,
    }
    if anomaly_row_count == 0:
        return {
            **base,
            "reason": "no_fdc_anomaly",
            "explanation": "FDC 이상이 없어 인과를 따질 근거 자체가 없습니다.",
        }
    if yield_delta_pp is None:
        return {
            **base,
            "reason": "no_baseline",
            "explanation": (
                "비교할 동일 공정 노드 평균이 없어 수율 저하 여부를 판정할 수 없습니다."
            ),
        }
    if yield_delta_pp > CAUSAL_YIELD_DEFICIT_PP:
        return {
            **base,
            "reason": "no_yield_deficit",
            "explanation": (
                f"FDC 이상 {anomaly_row_count}행이 있으나 다이 수율이 동일 공정 노드 평균 "
                f"대비 {yield_delta_pp:+.2f}%p로 저하가 없습니다. 이상이 수율로 이어졌다는 "
                "근거가 없습니다."
            ),
        }
    return {
        **base,
        "reason": "coincident_only",
        "explanation": (
            f"FDC 이상 {anomaly_row_count}행과 수율 저하 {yield_delta_pp:+.2f}%p가 함께 "
            "관측되지만, 비교군이 없는 단일 로트 관측이라 인과로 볼 수 없습니다. "
            "동시 발생까지가 이 데이터로 말할 수 있는 전부입니다."
        ),
    }


def check_source_consistency(tables: Tables, xlsx_path: Path) -> list[dict]:
    """CSV로 읽은 각 테이블이 번들 xlsx 시트와 행수가 일치하는지 검증하는 하네스 가드레일.

    dataset_2는 현재 CSV와 xlsx가 완전히 일치하지만, 실제 운영 데이터에서는
    소스가 갈라지는 게 흔하다 -- 이 검사가 실패하면 HITL로 넘겨야 할 신호다.
    """
    import openpyxl

    mismatches: list[dict] = []
    if not Path(xlsx_path).exists():
        return [{"table": "*", "issue": "xlsx_not_found", "path": str(xlsx_path)}]

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    for name in TABLE_FILES:
        if name not in wb.sheetnames:
            mismatches.append({"table": name, "issue": "sheet_missing"})
            continue
        ws = wb[name]
        # 값이 하나라도 있는 행만 센다. 실제 엑셀 파일은 데이터 아래에 서식만
        # 남은 빈 행을 수백 줄 끌고 다니는 일이 흔한데(dataset_2 번들 xlsx의
        # 차원 시트가 그렇다 — 5행짜리 시트가 1000행으로 선언되어 있다),
        # read_only 모드의 iter_rows는 그 빈 행까지 그대로 돌려준다. 값 검사
        # 없이 세면 데이터 행이 아니라 '행 슬롯'을 세게 되어, 소스가 멀쩡한데도
        # 항상 row_count_mismatch로 오탐한다.
        xlsx_row_count = sum(
            1 for row in ws.iter_rows(min_row=2)
            if any(cell.value is not None for cell in row)
        )
        csv_row_count = len(getattr(tables, name))
        if xlsx_row_count != csv_row_count:
            mismatches.append({
                "table": name,
                "issue": "row_count_mismatch",
                "csv_rows": csv_row_count,
                "xlsx_rows": xlsx_row_count,
            })
    return mismatches

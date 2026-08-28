"""가드레일 테스트 -- CSV/xlsx 소스 정합성, 통합 리포트 스키마, 그래프 결과 스키마.

positive case(현재 dataset_2처럼 완전히 일치)뿐 아니라, 가드레일이 실제로
불일치를 잡아내는 negative case까지 검증한다 -- 그래야 이 가드레일이
"항상 통과하는 장식"이 아니라는 걸 보증할 수 있다.
"""
from pathlib import Path

import openpyxl
import pytest
from openpyxl.styles import Border, Side

from insight_agent import domain
from insight_agent.harness.guardrails import validate_graph_result, validate_report

FIXTURES_DIR = Path(__file__).parent / "fixtures"

# fixtures의 실제 행수와 정확히 맞춘 값 (tests/test_domain.py의 row-count 테스트와 동일 기준)
ROW_COUNTS = {
    "dim_ai_automation_role": 2,
    "dim_process_design": 2,
    "dim_semicon_dx_architecture": 1,
    "fact_wafer_lot_yield": 4,
    "fact_fdc_chamber_sensor": 5,
    "fact_dx_smart_factory_kpi": 2,
}


def _build_xlsx(path: Path, row_counts: dict[str, int], skip_sheets: set[str] = frozenset()) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, n_rows in row_counts.items():
        if name in skip_sheets:
            continue
        ws = wb.create_sheet(name)
        ws.append(["col_a"])  # 헤더 1행
        for i in range(n_rows):
            ws.append([i])
    wb.save(path)


@pytest.fixture
def tables():
    return domain.load_tables(FIXTURES_DIR)


def test_source_consistency_passes_when_xlsx_matches(tables, tmp_path):
    xlsx_path = tmp_path / "matching.xlsx"
    _build_xlsx(xlsx_path, ROW_COUNTS)
    assert domain.check_source_consistency(tables, xlsx_path) == []


def test_source_consistency_detects_row_count_mismatch(tables, tmp_path):
    xlsx_path = tmp_path / "mismatched.xlsx"
    bad_counts = dict(ROW_COUNTS)
    bad_counts["fact_fdc_chamber_sensor"] = 3  # 실제 CSV는 5건인데 xlsx엔 3건만
    _build_xlsx(xlsx_path, bad_counts)

    mismatches = domain.check_source_consistency(tables, xlsx_path)

    assert len(mismatches) == 1
    assert mismatches[0] == {
        "table": "fact_fdc_chamber_sensor",
        "issue": "row_count_mismatch",
        "csv_rows": 5,
        "xlsx_rows": 3,
    }


def test_source_consistency_ignores_trailing_blank_rows(tables, tmp_path):
    """데이터 아래에 서식만 남은 빈 행이 수백 줄 붙어 있어도 데이터 행수만 세야 한다.

    실제 엑셀 파일(dataset_2 번들 xlsx의 차원 시트)이 정확히 이 모양이다 --
    5행짜리 시트가 1000행으로 선언되어 있다. 값 검사 없이 iter_rows를 세면
    소스가 멀쩡한데도 항상 row_count_mismatch로 오탐한다.
    """
    xlsx_path = tmp_path / "trailing_blank.xlsx"
    _build_xlsx(xlsx_path, ROW_COUNTS)

    # 값 없이 테두리 서식만 준 셀 -> 시트 dimension이 1000행까지 늘어난다
    wb = openpyxl.load_workbook(xlsx_path)
    for name in ROW_COUNTS:
        wb[name].cell(row=1000, column=1).border = Border(left=Side(style="thin"))
    wb.save(xlsx_path)

    assert domain.check_source_consistency(tables, xlsx_path) == []


def test_source_consistency_detects_missing_sheet(tables, tmp_path):
    xlsx_path = tmp_path / "missing_sheet.xlsx"
    _build_xlsx(xlsx_path, ROW_COUNTS, skip_sheets={"fact_dx_smart_factory_kpi"})

    mismatches = domain.check_source_consistency(tables, xlsx_path)

    assert {"table": "fact_dx_smart_factory_kpi", "issue": "sheet_missing"} in mismatches


def test_source_consistency_detects_missing_xlsx_file(tables, tmp_path):
    missing_path = tmp_path / "does_not_exist.xlsx"
    mismatches = domain.check_source_consistency(tables, missing_path)
    assert mismatches[0]["issue"] == "xlsx_not_found"


VALID_VERDICT = {
    "verdict": "UNKNOWN",
    "reason": "no_fdc_anomaly",
    "explanation": "FDC 이상이 없어 인과를 따질 근거 자체가 없습니다.",
}


def _report(**overrides):
    """필수 필드를 갖춘 최소 리포트. 검사하려는 항목만 덮어쓴다."""
    base = {
        "lot_id": "LOT-F001",
        "fdc_interlock_count": 0,
        "fdc_vm_error_anomaly_count": 0,
        "die_yield_pct": 80.0,
        "causal_verdict": dict(VALID_VERDICT),
    }
    base.update(overrides)
    return base


def test_validate_report_accepts_well_formed_report():
    validate_report(_report())


def test_validate_report_rejects_missing_fields():
    with pytest.raises(ValueError):
        validate_report({"lot_id": "LOT-F001"})


def test_validate_report_rejects_negative_interlock_count():
    # 다른 필드는 모두 갖춘 상태여야 '음수 거부'를 검증하는 테스트가 된다.
    with pytest.raises(ValueError, match="fdc_interlock_count"):
        validate_report(_report(fdc_interlock_count=-1))


def test_validate_report_requires_causal_verdict():
    """인과 판정 없이 리포트가 승인 단계로 넘어가면 안 된다."""
    report = _report()
    del report["causal_verdict"]
    with pytest.raises(ValueError, match="causal_verdict"):
        validate_report(report)


def test_validate_report_rejects_asserted_causation():
    """이 리포트는 인과를 확정할 수 없다 -- UNKNOWN 외의 판정은 가드레일이 막는다."""
    with pytest.raises(ValueError, match="must be one of"):
        validate_report(_report(causal_verdict={
            "verdict": "CONFIRMED",
            "reason": "fdc_anomaly",
            "explanation": "설비 이상이 수율을 떨어뜨렸다.",
        }))


def test_validate_report_rejects_unknown_without_a_reason():
    """사유 없는 UNKNOWN은 '모르겠다'가 아니라 판단 회피다."""
    with pytest.raises(ValueError, match="reason and explanation"):
        validate_report(_report(causal_verdict={
            "verdict": "UNKNOWN", "reason": "", "explanation": "",
        }))


def test_validate_graph_result_accepts_well_formed_result():
    validate_graph_result({"seed_nodes": [], "nodes": [], "facts": [], "context_text": ""})


def test_validate_graph_result_rejects_missing_fields():
    with pytest.raises(ValueError):
        validate_graph_result({"nodes": [], "facts": []})
